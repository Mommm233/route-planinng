import numpy as np
import matplotlib.pyplot as plt
import heapq
import math
import openpyxl
import time
from scipy.interpolate import UnivariateSpline

np.random.seed(0)
HEIGHT, WIDTH = 10, 10

fig, ax = plt.subplots()

map = np.array([[255 for _ in range(WIDTH)] for _ in range(HEIGHT)])

def plotPoint(_2DIndex, font):
    '''绘制文本'''
    row, col = _2DIndex
    ax.text(col, row, font, fontsize=20, fontweight='bold', color='white', ha='center', va='center')

def plotMap():
    ax.grid(True, color='black')
    ax.set_xticks(np.arange(-0.5, WIDTH, 1), np.arange(0, WIDTH + 1, 1), alpha=0)   # alpha=0隐藏刻度
    ax.set_yticks(np.arange(-0.5, HEIGHT, 1), np.arange(HEIGHT, -1, -1), alpha=0)
    ax.imshow(map, cmap='gray', vmin=0, vmax=255)
    plt.savefig('./image/aStar2.png')
    plt.show()


def to2DIndex(index):
    return (index // WIDTH, index % WIDTH)

def createObstacle(rate=0.1):
    '''障碍物''' 
    obstacle = np.random.choice(
        [_ for _ in range(HEIGHT * WIDTH)],
        size=int(HEIGHT * WIDTH * rate),
        replace=False
    )
    obstacle_row, obstacle_col = to2DIndex(obstacle)
    map[obstacle_row, obstacle_col] = 0
    map[3,7] = 0 # 手动添加一个


createObstacle()


start, end = (HEIGHT - 1, 0), (0, WIDTH - 1)
map[start], map[end] = 83, 69
plotPoint(start, 'S'), plotPoint(end, 'E')


directions = [
    (-1, 0),    # 上
    (1, 0),     # 下
    (0, -1),    # 左
    (0, 1),     # 右
    (-1, -1),   # 左上
    (-1, 1),    # 右上
    (1, -1),    # 左下
    (1, 1)      # 右下
]

class indexNode():
    def __init__(self, row=0, col=0):
        self.row = row
        self.col = col
        self.f = 0
        self.g = 0
        self.h = 0
        self.father = None
        # self.direction = (0, 1)

    def __lt__(self, a):
        return self.f < a.f
    
def getEstimate(cur, end):
    '''h(n)'''
    return math.sqrt((cur[0] - end[0]) ** 2 + (cur[1] - end[1]) ** 2)

def getGround(cur, next):
    '''g(n)'''
    return 1 if abs(cur[0] - next[0]) + abs(cur[1] - next[1]) == 1 else 1.414

def checkBoundary(next):
    row, col = next
    return row < 0 or row >= HEIGHT or col < 0 or col >= WIDTH

def checkObstacle(map, cur, next):
    '''安全性的检查障碍物'''
    if map[next] == 0: return True
    if abs(cur[0] - next[0]) + abs(cur[1] - next[1]) != 2:
        return False
    # 斜线
    diagonal1, diagonal2 = (cur[0], next[1]), (next[0], cur[1])
    if not checkBoundary(diagonal1) and not checkBoundary(diagonal2):
        return map[diagonal1] == 0 and map[diagonal2] == 0
    else: return False

def getNextNode(curNode, next, end):
    cur = (curNode.row, curNode.col)
    nextNode = indexNode(next[0], next[1])
    nextNode.g = curNode.g + getGround(cur, next)
    nextNode.h = getEstimate(next, end)
    nextNode.f = nextNode.g + nextNode.h
    nextNode.father = curNode
    return nextNode

def updateNextNode(curNode, nextNode):
    cur = (curNode.row, curNode.col)
    next = (nextNode.row, nextNode.col)
    nextNode.g = curNode.g + getGround(cur, next)
    nextNode.f = nextNode.g + nextNode.h
    nextNode.father = curNode

def pathRestore(forward_curNode, backward_curNode):
    pathNode = []
    while forward_curNode:
        pathNode.append(forward_curNode)
        forward_curNode = forward_curNode.father
    pathNode.reverse()
    while backward_curNode:
        pathNode.append(backward_curNode)
        backward_curNode = backward_curNode.father
    return pathNode

def extensionNode(map, open_list, open_dict, close_dict, curNode, goal):
    '''扩展节点'''
    cur = (curNode.row, curNode.col)
    for d in directions:
        next = (cur[0] + d[0], cur[1] + d[1])
        if checkBoundary(next) or checkObstacle(map, cur, next) or next in close_dict:
            continue    # 越界 障碍 已搜索
        if next in open_dict:
            nextNode = open_dict[next]
            if nextNode.g > curNode.g + getGround(cur, next):
                updateNextNode(curNode, nextNode)
                heapq.heapify(open_list)
        else:
            nextNode = getNextNode(curNode, next, goal)
            heapq.heappush(open_list, nextNode)
            open_dict.update({next : nextNode})
    del open_dict[cur]
    close_dict.update({cur : curNode})


def duplexAStar(map, start, end):
    '''双向A*'''
    forward_open_list, forward_open_dict = [], {}
    forward_close_dict = {}
    backward_open_list, backward_open_dict = [], {}
    backward_close_dict = {}

    startNode = indexNode(start[0], start[1])
    endNode = indexNode(end[0], end[1])
    startNode.f = startNode.h = \
        endNode.f = endNode.h = getEstimate(start, end)

    heapq.heapify(forward_open_list) # 堆化
    heapq.heappush(forward_open_list, startNode)
    forward_open_dict.update({start : startNode})

    heapq.heapify(backward_open_list) 
    heapq.heappush(backward_open_list, endNode)
    backward_open_dict.update({end : endNode})

    while forward_open_list and backward_open_list:
        forward_curNode = heapq.heappop(forward_open_list)
        backward_curNode = heapq.heappop(backward_open_list)
        forward_cur = (forward_curNode.row, forward_curNode.col)
        backward_cur = (backward_curNode.row, backward_curNode.col)
        if forward_cur == backward_cur:
            return pathRestore(forward_curNode, backward_curNode)
        if forward_cur in backward_close_dict:
            return pathRestore(forward_curNode, backward_close_dict[forward_cur])
        if backward_cur in forward_close_dict:
            return pathRestore(backward_curNode, forward_close_dict[backward_cur])
        extensionNode(map, forward_open_list, forward_open_dict, \
            forward_close_dict, forward_curNode, end)
        extensionNode(map, backward_open_list, backward_open_dict, \
            backward_close_dict, backward_curNode, start)

    return []

startTime = time.time()

pathNode = duplexAStar(map, start, end)

endTime = time.time()

def plotPath(pathNode):
    '''平滑化处理'''
    row_list, col_list, num_list = [], [], []
    st = set()
    for node in pathNode:
        if (node.row, node.col) not in st:
            row_list.append(node.row)
            col_list.append(node.col)
            num_list.append((node.f, node.g, node.h))
            st.add((node.row, node.col))
    # ax.plot(col_list, row_list, 'b')
    # 生成样条插值函数
    sorted_indices = np.argsort(col_list)
    print(sorted_indices)
    col_sorted = col_list[sorted_indices]
    row_sorted = row_list[sorted_indices]
    spline = UnivariateSpline(col_sorted, row_sorted, k=3)

    # 生成平滑后的数据
    col_smooth = np.linspace(col_sorted.min(), row_sorted.max(), 100)
    row_smooth = spline(col_smooth)
    ax.plot(col_smooth, row_smooth, 'r-', label='平滑后的数据')
    ax.legend()
    return row_list, col_list, num_list


row_list, col_list, num_list = plotPath(pathNode)


def creatResultXlsx(row_list, col_list, num_list):
    wb = openpyxl.Workbook()    # 创建一个新的Excel文件
    ws = wb.active  # 选择活动的工作表

    for i in range(1, len(row_list) + 1, 1):
        ws[f'A{i}'] =  f'({row_list[i - 1]}, {col_list[i - 1]})'
        ws[f'B{i}'] = num_list[i - 1][0]
        ws[f'C{i}'] = num_list[i - 1][1]
        ws[f'D{i}'] = num_list[i - 1][2]
    ws[f'A{len(row_list) + 1}'] = 'cost time'
    ws[f'B{len(row_list) + 1}'] = f'{endTime - startTime: .6f}s'
    wb.save("result2.xlsx")

creatResultXlsx(row_list, col_list, num_list)

plotMap()